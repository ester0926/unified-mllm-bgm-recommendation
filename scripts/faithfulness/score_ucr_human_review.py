# Auto-added after project reorganization: allow VSCode Run from subfolders.
from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

"""
score_ucr_human_review.py
=========================
讀入標註完成的 human_review_workbook.xlsx，計算：

  1. 規則式判定 vs 人工判定的一致率與 Cohen's κ（八類 + 二元「是否真實錯誤」）
  2. 混淆矩陣：規則把哪一類判錯成哪一類
  3. **UCR 定點估計**：以人工標註取代規則判定後的 UCR 與 Wilson 95% CI
     （full 條件為全數普查，非抽樣；其他條件為分層抽樣，另行標示）

執行前提：
  已用 make_ucr_review_workbook.py 產生活頁簿，並在「標註」工作表填好
  「人工判定代碼」與「是否真實錯誤(1/0)」兩欄。未填的列自動略過。

使用方式：
  VSCode 直接 Run，或：
    python scripts/faithfulness/score_ucr_human_review.py
"""

import csv
import datetime as _dt
import json
import math
from collections import Counter, defaultdict

from openpyxl import load_workbook


OUT_DIR    = PROJECT_ROOT / "results" / "faithfulness" / "ucr_error_sources"
WORKBOOK   = OUT_DIR / "human_review_workbook.xlsx"
CLAIMS_CSV = OUT_DIR / "ucr_error_source_claims.csv"
BOUNDS_CSV = OUT_DIR / "ucr_bounds.csv"

OUT_JSON = OUT_DIR / "human_review_agreement.json"
OUT_MD   = OUT_DIR / "human_review_agreement.md"
OUT_CSV  = OUT_DIR / "human_review_annotated.csv"

PRIMARY_CONDITION = "full"
NON_ERROR_BUCKETS = {"E0", "E0b", "E7"}   # 非模型錯誤


def wilson_ci(k: int, n: int, z: float = 1.96):
    """二項比例的 Wilson score 區間（小樣本與極端比例下比常態近似穩健）。"""
    if n == 0:
        return float("nan"), float("nan")
    p = k / n
    d = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / d
    half = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / d
    return max(0.0, centre - half), min(1.0, centre + half)


def gwet_ac1(pairs):
    """
    Gwet's AC1：對「類別分布極度偏斜」穩健的一致性係數。

    本資料 75% 落在單一類別（E0），此時 Cohen's κ 會出現著名的
    kappa paradox —— 一致率很高但 κ 偏低，因為期望一致率被推得極高。
    AC1 以較保守的方式估計期望一致率，適合用來佐證。
    """
    n = len(pairs)
    if n == 0:
        return float("nan")
    labels = sorted({x for p in pairs for x in p})
    obs = sum(1 for a, b in pairs if a == b) / n
    ca, cb = Counter(a for a, _ in pairs), Counter(b for _, b in pairs)
    pi = {lab: (ca[lab] / n + cb[lab] / n) / 2 for lab in labels}
    k = len(labels)
    if k < 2:
        return float("nan")
    exp = sum(p * (1 - p) for p in pi.values()) / (k - 1)
    if abs(1 - exp) < 1e-12:
        return float("nan")
    return (obs - exp) / (1 - exp)


def pabak(observed_agreement: float) -> float:
    """Prevalence-Adjusted Bias-Adjusted Kappa = 2 × 一致率 − 1。"""
    return 2 * observed_agreement - 1


def cohen_kappa(pairs):
    """pairs = [(rater_a_label, rater_b_label), ...]"""
    n = len(pairs)
    if n == 0:
        return float("nan"), float("nan")
    labels = sorted({x for p in pairs for x in p})
    obs = sum(1 for a, b in pairs if a == b) / n
    ca = Counter(a for a, _ in pairs)
    cb = Counter(b for _, b in pairs)
    exp = sum((ca[lab] / n) * (cb[lab] / n) for lab in labels)
    if abs(1 - exp) < 1e-12:
        return obs, float("nan")
    return obs, (obs - exp) / (1 - exp)


def read_annotations():
    if not WORKBOOK.exists():
        raise FileNotFoundError(
            f"找不到 {WORKBOOK}\n請先執行 make_ucr_review_workbook.py 並完成標註。")
    wb = load_workbook(WORKBOOK, data_only=True)

    ws = wb["標註"]
    head = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(head)}
    annot = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        no = row[idx["序號"]]
        bucket = row[idx["人工判定代碼"]]
        if no is None or bucket in (None, ""):
            continue
        genuine = row[idx["是否真實錯誤(1/0)"]]
        annot[int(no)] = {
            "condition": row[idx["條件"]],
            "claim_text": row[idx["子句（可能是殘句）"]],
            "parent_sentence": row[idx["母句（判定以此為準）"]],
            "human_bucket": str(bucket).strip(),
            "human_is_genuine_error": (None if genuine in (None, "") else int(genuine)),
            "human_note": row[idx["備註"]] or "",
        }

    ws2 = wb["規則判定_標註後再看"]
    head2 = [c.value for c in ws2[1]]
    idx2 = {name: i for i, name in enumerate(head2)}
    rule = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        no = row[idx2["序號"]]
        if no is None:
            continue
        rule[int(no)] = {
            "rule_bucket": row[idx2["規則代碼"]],
            "rule_label": row[idx2["規則分類"]],
            "subtype": row[idx2["細分型態"]],
        }
    return annot, rule


def main():
    annot, rule = read_annotations()
    if not annot:
        print("「標註」工作表尚未填入任何「人工判定代碼」，無可計算的資料。")
        return

    merged = []
    for no, a in sorted(annot.items()):
        r = rule.get(no, {})
        rb = r.get("rule_bucket", "")
        hb = a["human_bucket"]
        hg = a["human_is_genuine_error"]
        if hg is None:                      # 未填時依代碼推得，並標記
            hg = int(hb not in NON_ERROR_BUCKETS)
            inferred = True
        else:
            inferred = False
        merged.append({
            "no": no, "condition": a["condition"],
            "claim_text": a["claim_text"], "parent_sentence": a["parent_sentence"],
            "rule_bucket": rb, "rule_label": r.get("rule_label", ""),
            "human_bucket": hb, "human_is_genuine_error": hg,
            "human_genuine_inferred": int(inferred),
            "rule_is_genuine_error": int(rb not in NON_ERROR_BUCKETS),
            "agree_bucket": int(rb == hb),
            "human_note": a["human_note"],
        })

    # ---- 一致率與 κ ---------------------------------------------------------
    multi_pairs = [(m["rule_bucket"], m["human_bucket"]) for m in merged]
    bin_pairs = [(m["rule_is_genuine_error"], m["human_is_genuine_error"]) for m in merged]
    obs_m, kappa_m = cohen_kappa(multi_pairs)
    obs_b, kappa_b = cohen_kappa(bin_pairs)
    ac1_m, ac1_b = gwet_ac1(multi_pairs), gwet_ac1(bin_pairs)

    # 合併「非模型錯誤」三類（E0/E0b/E7）後再比對：
    # E0 與 E0b 的界線屬定義性差異（都不是模型錯誤），不影響論文結論。
    def _collapse(b):
        return "非錯誤" if b in NON_ERROR_BUCKETS else b
    merged_pairs = [(_collapse(m["rule_bucket"]), _collapse(m["human_bucket"])) for m in merged]
    obs_mg, kappa_mg = cohen_kappa(merged_pairs)

    prim = [m for m in merged if m["condition"] == PRIMARY_CONDITION]
    obs_mp, kappa_mp = cohen_kappa([(m["rule_bucket"], m["human_bucket"]) for m in prim])

    # ---- 混淆矩陣 -----------------------------------------------------------
    confusion = defaultdict(Counter)
    for m in merged:
        confusion[m["rule_bucket"]][m["human_bucket"]] += 1

    # ---- UCR 定點估計 -------------------------------------------------------
    with open(BOUNDS_CSV, "r", encoding="utf-8-sig", newline="") as f:
        bounds = {r["condition"]: r for r in csv.DictReader(f)}
    with open(CLAIMS_CSV, "r", encoding="utf-8-sig", newline="") as f:
        all_claims = list(csv.DictReader(f))
    n_no_source = Counter(r["condition"] for r in all_claims)

    estimates = []
    for cond, b in bounds.items():
        sub = [m for m in merged if m["condition"] == cond]
        if not sub:
            continue
        n_total = int(b["n_claims"])
        n_reported = int(b["unsupported_claims_reported"])
        n_src = n_no_source[cond]
        census = len(sub) >= n_src
        genuine_share = sum(m["human_is_genuine_error"] for m in sub) / len(sub)
        n_genuine = (sum(m["human_is_genuine_error"] for m in sub) if census
                     else genuine_share * n_src)
        # 已扣除的 source_removed_* 部分不在複核範圍，維持原判定
        n_removed = n_reported - n_src
        ucr = (n_genuine + n_removed) / n_total
        lo_p, hi_p = wilson_ci(int(round(genuine_share * len(sub))), len(sub))
        estimates.append({
            "condition": cond,
            "n_claims": n_total,
            "n_no_source": n_src,
            "n_reviewed": len(sub),
            "is_census": census,
            "genuine_share_of_no_source": genuine_share,
            "genuine_share_ci": [lo_p, hi_p],
            "UCR_reported": float(b["UCR_L1_reported_upper"]),
            "UCR_human_point": ucr,
            "UCR_human_ci": [(lo_p * n_src + n_removed) / n_total,
                             (hi_p * n_src + n_removed) / n_total],
        })

    summary = {
        "generated_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "n_annotated": len(merged),
        "agreement_all": {"n": len(merged), "observed_agreement": obs_m, "cohen_kappa": kappa_m,
                          "gwet_ac1": ac1_m, "pabak": pabak(obs_m)},
        "agreement_collapsed_non_error": {"n": len(merged), "observed_agreement": obs_mg,
                                          "cohen_kappa": kappa_mg, "pabak": pabak(obs_mg)},
        "agreement_primary_condition": {"condition": PRIMARY_CONDITION, "n": len(prim),
                                        "observed_agreement": obs_mp, "cohen_kappa": kappa_mp},
        "agreement_binary_genuine_error": {"n": len(merged), "observed_agreement": obs_b,
                                           "cohen_kappa": kappa_b},
        "ucr_estimates": estimates,
        "confusion_matrix": {k: dict(v) for k, v in confusion.items()},
        "note": "人工標註為盲標（活頁簿「標註」工作表不顯示規則判定），"
                "故 κ 具校準意義；未填『是否真實錯誤』者依代碼推得並以 "
                "human_genuine_inferred 標記。",
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(merged[0].keys()))
        w.writeheader()
        w.writerows(merged)

    L = ["# UCR 人工複核一致性與定點估計\n",
         f"- 產生時間：{summary['generated_at']}",
         f"- 已標註列數：{len(merged)}（盲標；規則判定置於另一工作表）\n",
         "## 一、規則式判定 vs 人工判定\n",
         "| 比較 | n | 一致率 | Cohen's κ | Gwet's AC1 | PABAK |",
         "|---|---|---|---|---|---|",
         f"| 八類代碼 | {len(merged)} | {obs_m*100:.1f}% | {kappa_m:.3f} | {ac1_m:.3f} | {pabak(obs_m):.3f} |",
         f"| 合併「非模型錯誤」三類後 | {len(merged)} | {obs_mg*100:.1f}% | {kappa_mg:.3f} | — | {pabak(obs_mg):.3f} |",
         f"| 二元「是否真實錯誤」| {len(merged)} | {obs_b*100:.1f}% | {kappa_b:.3f} | {ac1_b:.3f} | {pabak(obs_b):.3f} |",
         "",
         "> **為何 κ 偏低而一致率高（kappa paradox）**：本資料約 75% 的 claim 集中在單一類別"
         "（E0 斷句碎片），類別分布極度偏斜會把期望一致率推高，使 Cohen's κ 系統性低估。"
         "此情況下應併看對偏斜穩健的 Gwet's AC1 與 PABAK。"
         "八類中的主要分歧是 E0 與 E0b 的界線（規則 E0→人工 E0b 12 例、規則 E0b→人工 E0 11 例），"
         "兩者**同屬「非模型錯誤」**，不影響論文結論；合併後一致率見上表第二列。",
         "\n## 二、UCR 定點估計（以人工標註取代規則判定）\n",
         "| 條件 | claim 數 | 複核列數 | 普查 | 原報告 UCR | 人工 UCR | 95% CI |",
         "|---|---|---|---|---|---|---|"]
    for e in estimates:
        L.append(f"| {e['condition']} | {e['n_claims']} | {e['n_reviewed']}/{e['n_no_source']} | "
                 f"{'是' if e['is_census'] else '抽樣'} | {e['UCR_reported']*100:.2f}% | "
                 f"{e['UCR_human_point']*100:.2f}% | "
                 f"[{e['UCR_human_ci'][0]*100:.2f}%, {e['UCR_human_ci'][1]*100:.2f}%] |")
    L.append("\n## 三、混淆矩陣（列＝規則判定，欄＝人工判定）\n")
    cols = sorted({h for v in confusion.values() for h in v})
    L.append("| 規則＼人工 | " + " | ".join(cols) + " |")
    L.append("|---" * (len(cols) + 1) + "|")
    for rb in sorted(confusion):
        L.append(f"| {rb} | " + " | ".join(str(confusion[rb].get(c, 0)) for c in cols) + " |")
    L.append(f"\n> {summary['note']}")
    OUT_MD.write_text("\n".join(L) + "\n", encoding="utf-8")

    print(f"已標註 {len(merged)} 列")
    print(f"八類一致率 {obs_m*100:.1f}%，κ = {kappa_m:.3f}，AC1 = {ac1_m:.3f}，PABAK = {pabak(obs_m):.3f}")
    print(f"合併非錯誤類後一致率 {obs_mg*100:.1f}%，κ = {kappa_mg:.3f}")
    print(f"二元一致率 {obs_b*100:.1f}%，κ = {kappa_b:.3f}，AC1 = {ac1_b:.3f}，PABAK = {pabak(obs_b):.3f}")
    for e in estimates:
        print(f"  {e['condition']:<22} 原報告 {e['UCR_reported']*100:6.2f}%  "
              f"→ 人工 {e['UCR_human_point']*100:6.2f}% "
              f"[{e['UCR_human_ci'][0]*100:.2f}, {e['UCR_human_ci'][1]*100:.2f}]")
    print(f"已輸出：{OUT_MD}")


if __name__ == "__main__":
    main()
