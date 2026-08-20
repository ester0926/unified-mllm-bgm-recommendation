"""
用途：整理人工複查用資料與 UCR 錯誤來源分析結果。
輸入：主評估輸出的推薦解釋、metadata、counterfactual 或人工複查檔。
輸出：claim 標註、faithfulness 指標、UCR 摘要或人工檢查表。
執行：通常需先完成主評估或 Top-1 生成，再執行本檔。
"""

from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


import csv
import datetime as _dt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


IN_CSV = PROJECT_ROOT / "results" / "faithfulness" / "ucr_error_sources" / "human_review_template.csv"
OUT_XLSX = PROJECT_ROOT / "results" / "faithfulness" / "ucr_error_sources" / "human_review_workbook.xlsx"

PRIMARY_CONDITION = "full"

# 安全開關：既有活頁簿若已含人工標註，預設拒絕覆蓋（避免重跑洗掉標註成果）。
FORCE_OVERWRITE = "--force" in sys.argv

BUCKET_DEFS = [
    ("E7", "生成崩潰（退化輸出）",
     "整段生成是重複、亂碼或空輸出，根本不成句子。此時談「主張有無證據」沒有意義。"),
    ("E0", "斷句碎片（標註流程假影）",
     "子句被切在 and/逗號中間而不成主張，但**母句**本身可以歸屬到某個來源"
     "（音訊、元資料、影片、提示詞、偏好、一般推薦語）。這不是模型的錯。"),
    ("E0b", "規則漏判但有證據支持",
     "母句是完整主張，且內容可由參考文本或 musicnn 標籤佐證，只是規則字典沒抓到。也不是模型的錯。"),
    ("E1", "偏好不支持",
     "母句主張使用者的偏好（你通常喜歡…／符合你的品味），但沒有任何偏好證據可佐證。"),
    ("E2", "音樂元資料不支持",
     "母句主張曲風、專輯、演出者等音樂事實，但參考文本與標籤都查不到支持。"
     "注意：musicnn 標籤有雜訊，「查不到」不等於「講錯」。"),
    ("E3", "影片內容不支持",
     "母句主張具體影片情境（婚禮／旅遊／教學…），但模型輸入只有 CLIP 影像特徵，無從佐證。"),
    ("E4", "一般性空泛描述／主觀陳述",
     "母句是泛用效用宣稱（能為你的影片加分）或主觀感受（很美、很懷舊），本質上不可驗證。"),
    ("E5", "多來源無法區分",
     "同一句同時混合音樂屬性與影片效用等多個來源，無法歸屬到單一來源。"),
    ("E6", "可能幻覺",
     "母句提出**具體且應該可以查證**的內容（如專輯／mixtape 名稱），但所有證據來源都查無此事。"),
]
BUCKET_CODES = [b[0] for b in BUCKET_DEFS]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TIER_FILL = {"A": PatternFill("solid", fgColor="C6E0B4"),
             "B": PatternFill("solid", fgColor="FFE699"),
             "C": PatternFill("solid", fgColor="F2F2F2")}


def read_rows(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def tier_of(row):
    if row["condition"] == PRIMARY_CONDITION:
        return "A"
    return "B" if row.get("needs_human_review") == "1" else "C"


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def count_existing_annotations() -> int:
    """回傳既有活頁簿中已填寫「人工判定代碼」的列數；檔案不存在或無法讀取時回傳 0。"""
    if not OUT_XLSX.exists():
        return 0
    try:
        from openpyxl import load_workbook
        wb = load_workbook(OUT_XLSX, data_only=True)
        if "標註" not in wb.sheetnames:
            return 0
        ws = wb["標註"]
        header = [c.value for c in ws[1]]
        if "人工判定代碼" not in header:
            return 0
        col = header.index("人工判定代碼")
        return sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                   if row[col] not in (None, ""))
    except Exception:                      # 檔案毀損或被鎖住時不阻擋，交由後續寫入報錯
        return 0


def build():
    # 防呆：若既有活頁簿已有人工標註，重新產生會把標註全部洗掉。
    n_existing = count_existing_annotations()
    if n_existing and not FORCE_OVERWRITE:
        raise SystemExit(
            f"中止：{OUT_XLSX.name} 已有 {n_existing} 列人工標註，重新產生會全部覆蓋。\n"
            f"若確定要重建空白活頁簿，請先備份該檔，再將本檔的 FORCE_OVERWRITE 設為 True\n"
            f"（或在命令列加上 --force）。"
        )

    rows = read_rows(IN_CSV)
    for i, r in enumerate(rows, start=1):
        r["_tier"] = tier_of(r)
    rows.sort(key=lambda r: (r["_tier"], r["condition"], int(r["sample_idx"]), int(r["claim_id"])))
    for i, r in enumerate(rows, start=1):
        r["_no"] = i

    wb = Workbook()

    # ---------------- 工作表 1：標註（盲標）--------------------------------
    ws = wb.active
    ws.title = "標註"
    headers = ["序號", "優先", "條件", "子句（可能是殘句）", "母句（判定以此為準）",
               "人工判定代碼", "是否真實錯誤(1/0)", "備註", "生成全文（需要時參考）"]
    ws.append(headers)
    for r in rows:
        ws.append([r["_no"], r["_tier"], r["condition"], r["claim_text"],
                   r["parent_sentence"], "", "", "", r["generated_text"]])
    style_header(ws, len(headers))

    widths = [6, 6, 20, 46, 52, 14, 14, 26, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_cells[1].fill = TIER_FILL[row_cells[1].value]
        for idx in (5, 6, 7):          # 人工判定代碼 / 是否真實錯誤 / 備註
            row_cells[idx].fill = INPUT_FILL

    dv_bucket = DataValidation(type="list", formula1='"' + ",".join(BUCKET_CODES) + '"',
                               allow_blank=True, showDropDown=False)
    dv_bucket.error = "請從下拉選單選擇 E0/E0b/E1…E7"
    dv_bucket.errorTitle = "代碼不正確"
    ws.add_data_validation(dv_bucket)
    dv_bucket.add(f"F2:F{ws.max_row}")

    dv_bin = DataValidation(type="list", formula1='"1,0"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_bin)
    dv_bin.add(f"G2:G{ws.max_row}")

    # ---------------- 工作表 2：規則判定（標註後再看）-----------------------
    ws2 = wb.create_sheet("規則判定_標註後再看")
    headers2 = ["序號", "條件", "規則代碼", "規則分類", "細分型態", "判定依據",
                "曲風詞", "曲風驗證", "查無支持之曲風", "引號名稱", "名稱驗證", "規則標記需複核"]
    ws2.append(headers2)
    for r in rows:
        ws2.append([r["_no"], r["condition"], r["bucket"].split("_")[0], r["bucket_label"],
                    r["subtype"], r["evidence_note"], r["genre_terms"], r["genre_verification"],
                    r["genre_missing"], r["quoted_names"], r["quoted_verification"],
                    r["needs_human_review"]])
    style_header(ws2, len(headers2))
    for i, w in enumerate([6, 20, 10, 26, 30, 56, 20, 18, 18, 22, 14, 12], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row_cells in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ---------------- 工作表 3：標註準則 ------------------------------------
    ws3 = wb.create_sheet("標註準則")
    ws3.append(["UCR 錯誤來源人工標註準則"])
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.append([])
    ws3.append(["判定單位",
                "以「母句」為準。子句若被切在 and/逗號中間而不成主張，"
                "請看母句的完整意思再判定。"])
    ws3.append(["判定順序", "由上而下第一個成立者即為答案，每列只填一個代碼。"])
    ws3.append([])
    ws3.append(["代碼", "分類", "判定說明"])
    for code, label, desc in BUCKET_DEFS:
        ws3.append([code, label, desc])
    ws3.append([])
    ws3.append(["是否真實錯誤(1/0)",
                "填 1 = 這確實是「模型講了沒有證據支持的話」；"
                "填 0 = 這不是模型的錯（斷句造成、規則沒抓到、或生成已崩潰）。"])
    ws3.append(["", "對照關係：E0 / E0b / E7 一律填 0；E1–E6 一律填 1。"
                    "若你認為某列不符合此對照，請填你認為正確的值並在備註說明。"])
    ws3.append([])
    ws3.append(["重要提醒",
                "標註完成前請勿翻閱「規則判定_標註後再看」工作表，"
                "否則會產生錨定效應，使 Cohen's κ 失去校準意義。"])
    ws3.append(["優先順序",
                "A = full 條件（論文報告的 UCR 13.12% 即此條件，143 列，務必全標）；"
                "B = 其他條件中規則標記需複核者；C = 其餘抽樣列。"])
    ws3.append(["產生時間", _dt.datetime.now().isoformat(timespec="seconds")])
    for c, w in zip("ABC", (22, 30, 96)):
        ws3.column_dimensions[c].width = w
    for row_cells in ws3.iter_rows(min_row=1, max_row=ws3.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.cell(row=6, column=1).font = Font(bold=True)
    ws3.cell(row=6, column=2).font = Font(bold=True)
    ws3.cell(row=6, column=3).font = Font(bold=True)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    tiers = {}
    for r in rows:
        tiers[r["_tier"]] = tiers.get(r["_tier"], 0) + 1
    print(f"已輸出：{OUT_XLSX}")
    print(f"總列數 {len(rows)}｜A={tiers.get('A', 0)} B={tiers.get('B', 0)} C={tiers.get('C', 0)}")


if __name__ == "__main__":
    build()
