"""
用途：分析影片群集與分層抽樣設定。
輸入：既有實驗輸出、metadata、評估 CSV 或分析用中間檔。
輸出：論文分析用表格、圖表、摘要 JSON/CSV 或檢查清單。
執行：請先確認前一階段輸出檔已存在，再從 repo 根目錄執行。
"""

from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


import csv
import datetime as _dt
import json
import re
from collections import Counter, defaultdict

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RESULTS_DIR = PROJECT_ROOT / "results"
CACHE_DIR   = PROJECT_ROOT / "cache"
OUT_DIR     = RESULTS_DIR / "analysis" / "video_clusters"

FEATURE_CACHE   = CACHE_DIR / "test_video_features.npz"
ASSIGNMENTS_CSV = OUT_DIR / "video_cluster_assignments.csv"

YOUTUBE_METADATA = Path(
    r"data/user_profiling/music_metadata_simple\youtube_metadata.jsonl")
MUSIC_METADATA = Path(
    r"data/user_profiling/music_metadata_simple\music_metadata_enriched.json")

K_REPORT = [3, 4]
N_EXEMPLAR = 12          # 每叢集的質心代表樣本數
N_DISTINCTIVE = 12       # 每叢集的判別性樣本數
MIN_COUNT = 15           # 富集比計算的最低出現次數（避免長尾雜訊）
TOP_TERMS = 12

STOPWORDS = {
    "the", "a", "an", "of", "and", "or", "in", "on", "for", "to", "with", "by",
    "feat", "ft", "from", "at", "is", "it", "my", "you", "your", "me", "i",
    "official", "video", "music", "hd", "new", "full", "com", "www", "youtube",
    "mv", "ver", "vol", "part", "no", "de", "la", "el", "x", "s", "t",
}

TITLE_TOKEN_RE = re.compile(r"[A-Za-z][A-Za-z'-]{1,}")


def read_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_metadata():
    yt = {}
    if YOUTUBE_METADATA.exists():
        with open(YOUTUBE_METADATA, "r", encoding="utf-8") as f:
            for line in f:
                d = json.loads(line)
                yt[d["music_id"]] = d
    md = {}
    if MUSIC_METADATA.exists():
        md = json.loads(MUSIC_METADATA.read_text(encoding="utf-8"))
    return yt, md


def title_tokens(title: str) -> set:
    return {w.lower() for w in TITLE_TOKEN_RE.findall(title or "")
            if w.lower() not in STOPWORDS and len(w) > 2}


def enrichment(doc_terms, in_cluster_mask, min_count=MIN_COUNT, top=TOP_TERMS):
    """
    回傳 [(term, rate_in, rate_out, ratio, n_in)]，依 ratio 由大到小。
    ratio = 本叢集出現率 ÷ 其他叢集出現率（加 1 平滑避免除以 0）。
    """
    n_in = int(in_cluster_mask.sum())
    n_out = int((~in_cluster_mask).sum())
    if n_in == 0 or n_out == 0:
        return []
    cnt_in, cnt_out = Counter(), Counter()
    for i, terms in enumerate(doc_terms):
        target = cnt_in if in_cluster_mask[i] else cnt_out
        for t in terms:
            target[t] += 1
    out = []
    for t, c in cnt_in.items():
        if c < min_count:
            continue
        r_in = c / n_in
        r_out = (cnt_out.get(t, 0) + 1) / (n_out + 1)
        out.append((t, r_in, r_out, r_in / r_out, c))
    out.sort(key=lambda x: -x[3])
    return out[:top]


def main():
    if not FEATURE_CACHE.exists():
        raise FileNotFoundError(
            f"找不到特徵快取 {FEATURE_CACHE}\n請先執行 video_cluster_stratification.py。")
    if not ASSIGNMENTS_CSV.exists():
        raise FileNotFoundError(f"找不到叢集結果 {ASSIGNMENTS_CSV}")

    feats = np.load(FEATURE_CACHE, allow_pickle=True)["features"]
    rows = read_csv(ASSIGNMENTS_CSV)
    if len(rows) != feats.shape[0]:
        raise ValueError("叢集結果與特徵快取筆數不符，請重跑 video_cluster_stratification.py")

    yt, md = load_metadata()
    video_ids = [r["video_id"] for r in rows]
    titles = [yt.get(v, {}).get("title", "") for v in video_ids]
    doc_title_terms = [title_tokens(t) for t in titles]
    doc_tag_terms = []
    for v in video_ids:
        e = md.get(v, {})
        terms = {str(e.get("genre", "")).lower()} if e.get("genre") else set()
        terms |= {str(t).lower() for t in (e.get("tags") or [])}
        doc_tag_terms.append({t for t in terms if t})

    durations = np.array([yt.get(v, {}).get("duration") or np.nan for v in video_ids], dtype=float)
    views = np.array([yt.get(v, {}).get("view_count") or np.nan for v in video_ids], dtype=float)

    profile = {"generated_at": _dt.datetime.now().isoformat(timespec="seconds"), "clusters": {}}
    md_lines = ["# 叢集輪廓卡（B4 人工命名輔助）\n",
                f"- 產生時間：{profile['generated_at']}",
                "- 富集比 = 該詞在本叢集的出現率 ÷ 在其他叢集的出現率；**> 1.5 才算特徵詞**",
                "- ⚠ B4 的輪廓係數僅 0.051–0.073，代表叢集結構本就很弱。"
                "若下列特徵詞的富集比多半接近 1，就是「這批影片切不出內容類型」的直接證據，"
                "請據實記錄，不要勉強命名。\n"]

    wb = Workbook()
    wb.remove(wb.active)

    for k in K_REPORT:
        labels = np.array([int(r[f"cluster_k{k}"]) for r in rows])
        centroids = np.vstack([feats[labels == c].mean(axis=0) for c in range(k)])
        centroids /= np.linalg.norm(centroids, axis=1, keepdims=True)
        sims = feats @ centroids.T                       # 特徵已 L2 正規化 → 內積即餘弦

        md_lines.append(f"\n## k = {k}\n")
        for c in range(k):
            mask = labels == c
            n_c = int(mask.sum())

            top_titles = enrichment(doc_title_terms, mask)
            top_tags = enrichment(doc_tag_terms, mask)

            own = sims[:, c].copy()
            other = np.delete(sims, c, axis=1).max(axis=1)
            margin = own - other

            idx_c = np.flatnonzero(mask)
            exemplars = idx_c[np.argsort(-own[idx_c])][:N_EXEMPLAR]
            distinctive = idx_c[np.argsort(-margin[idx_c])][:N_DISTINCTIVE]

            info = {
                "k": k, "cluster": c, "n": n_c, "share": n_c / len(rows),
                "median_duration_sec": float(np.nanmedian(durations[mask])),
                "median_view_count": float(np.nanmedian(views[mask])),
                "top_title_terms": [{"term": t, "ratio": round(r, 2), "n": n}
                                    for t, _, _, r, n in top_titles],
                "top_tag_terms": [{"term": t, "ratio": round(r, 2), "n": n}
                                  for t, _, _, r, n in top_tags],
                "max_title_ratio": round(max([r for *_, r, _ in top_titles], default=0), 2),
                "max_tag_ratio": round(max([r for *_, r, _ in top_tags], default=0), 2),
            }
            profile["clusters"][f"k{k}_c{c}"] = info

            md_lines.append(f"### k={k} 叢集 {c}（n={n_c}，{n_c/len(rows)*100:.1f}%）\n")
            md_lines.append(f"- 影片長度中位數：{info['median_duration_sec']:.0f} 秒"
                            f"　觀看數中位數：{info['median_view_count']:,.0f}")
            md_lines.append(f"- **標題特徵詞**（富集比最高前 {TOP_TERMS}）：" +
                            ("、".join(f"`{t}`×{r:.2f}" for t, _, _, r, _ in top_titles)
                             if top_titles else "（無達門檻者）"))
            md_lines.append(f"- **曲風／標籤特徵詞**：" +
                            ("、".join(f"`{t}`×{r:.2f}" for t, _, _, r, _ in top_tags)
                             if top_tags else "（無達門檻者）"))
            md_lines.append("")

            # ---- Excel 分頁 ----
            ws = wb.create_sheet(f"k{k}_叢集{c}")
            ws.append([f"k={k} 叢集 {c}｜n={n_c}（{n_c/len(rows)*100:.1f}%）"
                       f"｜長度中位數 {info['median_duration_sec']:.0f}s"
                       f"｜觀看數中位數 {info['median_view_count']:,.0f}"])
            ws["A1"].font = Font(bold=True, size=12)
            ws.append(["標題特徵詞（富集比）",
                       "、".join(f"{t} ×{r:.2f}" for t, _, _, r, _ in top_titles) or "（無）"])
            ws.append(["曲風／標籤特徵詞",
                       "、".join(f"{t} ×{r:.2f}" for t, _, _, r, _ in top_tags) or "（無）"])
            ws.append([])
            ws.append(["★ 請在此填寫叢集名稱 →", ""])
            ws.append(["★ 若無法區分，請填「無法區分」並在此說明 →", ""])
            ws.append([])
            ws.append(["類型", "序", "影片標題", "演出者", "長度(秒)", "觀看數", "YouTube 連結"])
            for label, idxs in (("判別性樣本（優先看）", distinctive), ("質心代表樣本", exemplars)):
                for rank, i in enumerate(idxs, start=1):
                    v = video_ids[i]
                    m = yt.get(v, {})
                    ws.append([label, rank, m.get("title", ""), m.get("artist", ""),
                               m.get("duration", ""), m.get("view_count", ""),
                               m.get("youtube_url", f"https://www.youtube.com/watch?v={v}")])
            hdr = 8
            for col, w in zip("ABCDEFG", (22, 5, 58, 26, 10, 12, 46)):
                ws.column_dimensions[col].width = w
            for cell in ws[hdr]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F3864")
            for r_i in range(hdr + 1, ws.max_row + 1):
                ws.cell(row=r_i, column=7).hyperlink = ws.cell(row=r_i, column=7).value
                ws.cell(row=r_i, column=7).font = Font(color="0563C1", underline="single")
                for c_i in range(1, 8):
                    ws.cell(row=r_i, column=c_i).alignment = Alignment(vertical="top",
                                                                       wrap_text=True)
            for r_i in (5, 6):
                ws.cell(row=r_i, column=1).font = Font(bold=True)
                ws.cell(row=r_i, column=2).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.freeze_panes = f"A{hdr + 1}"

    md_lines.append("\n## 命名作業建議順序\n")
    md_lines.append("1. 先讀上面的輪廓卡：若某叢集的特徵詞富集比都在 1.5 以下，"
                    "代表它與其他叢集在標題與曲風上並無實質差異。")
    md_lines.append("2. 開啟 `cluster_naming_worksheet.xlsx`，每個叢集一張分頁。"
                    "**先看「判別性樣本」的 12 支**（最能凸顯叢集差異），必要時再看質心代表樣本。")
    md_lines.append("3. 在每張分頁上方的黃色欄位填入叢集名稱；"
                    "若看完仍分不出來，請填「無法區分」並簡述你看到什麼。")
    md_lines.append("4. 只需完成 k=3 的三張分頁即可下結論；k=4 供交叉確認，時間不足可略過。")

    (OUT_DIR / "cluster_profile.md").write_text("\n".join(md_lines) + "\n", encoding="utf-8")
    (OUT_DIR / "cluster_profile.json").write_text(
        json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    xlsx = OUT_DIR / "cluster_naming_worksheet.xlsx"
    wb.save(xlsx)

    print(f"已輸出：{OUT_DIR / 'cluster_profile.md'}")
    print(f"已輸出：{xlsx}")
    for key, info in profile["clusters"].items():
        print(f"  {key}: n={info['n']:5d} 標題最高富集比={info['max_title_ratio']:.2f} "
              f"標籤最高富集比={info['max_tag_ratio']:.2f}")


if __name__ == "__main__":
    main()
