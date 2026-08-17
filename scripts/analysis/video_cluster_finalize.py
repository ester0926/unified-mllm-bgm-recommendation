# Auto-added after project reorganization: allow VSCode Run from subfolders.
from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

"""
video_cluster_finalize.py
=========================
B4 收尾：把人工命名結果併回叢集分析，產出論文可直接引用的表格。

輸入：
  cluster_naming_worksheet.xlsx      人工填寫的叢集名稱（每分頁 B5 儲存格）
  video_cluster_metrics.csv          B4 的分層指標
  video_cluster_gain.csv             逐叢集 LTP 增益 + 拔靴 CI
  video_cluster_homogeneity.csv      增益同質性檢定（含置換檢定 p）
  video_cluster_quality.csv          各 k 的輪廓係數
  video_cluster_assignments.csv      逐樣本叢集標籤
  cluster_profile.json               特徵詞富集比

輸出（results/analysis/video_clusters/）：
  video_cluster_named_metrics.csv    含叢集名稱的完整指標表
  video_cluster_assignments_named.csv 逐樣本標籤 + 名稱（供後續 B6 使用）
  B4_論文表格.md                      §4.7 可直接引用的表格與建議敘述

使用方式：
  python scripts/analysis/video_cluster_finalize.py
"""

import csv
import datetime as _dt
import json

from openpyxl import load_workbook


OUT_DIR = PROJECT_ROOT / "results" / "analysis" / "video_clusters"
WORKBOOK = OUT_DIR / "cluster_naming_worksheet.xlsx"

NAME_CELL = (5, 2)     # B5：叢集名稱
NOTE_CELL = (6, 2)     # B6：無法區分時的說明


def read_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows):
    if not rows:
        return
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def load_names() -> dict:
    """回傳 {(k, cluster): {'name':…, 'note':…}}。"""
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"找不到命名工作表：{WORKBOOK}\n請先執行 video_cluster_profile.py 並完成命名。")
    wb = load_workbook(WORKBOOK, data_only=True)
    names = {}
    for sheet in wb.sheetnames:
        # 分頁名格式：k{K}_叢集{C}
        try:
            head, tail = sheet.split("_")
            k = int(head.lstrip("k"))
            c = int(tail.replace("叢集", ""))
        except (ValueError, IndexError):
            continue
        ws = wb[sheet]
        names[(k, c)] = {
            "name": (ws.cell(*NAME_CELL).value or "").strip(),
            "note": (ws.cell(*NOTE_CELL).value or "").strip(),
        }
    missing = [key for key, v in names.items() if not v["name"]]
    if missing:
        raise SystemExit(
            "以下叢集尚未填寫名稱，請完成後再執行：" +
            "、".join(f"k={k} 叢集{c}" for k, c in sorted(missing)))
    return names


def main():
    names = load_names()
    metrics = read_csv(OUT_DIR / "video_cluster_metrics.csv")
    gains = read_csv(OUT_DIR / "video_cluster_gain.csv")
    homo = read_csv(OUT_DIR / "video_cluster_homogeneity.csv")
    quality = read_csv(OUT_DIR / "video_cluster_quality.csv")
    assignments = read_csv(OUT_DIR / "video_cluster_assignments.csv")
    profile = json.loads((OUT_DIR / "cluster_profile.json").read_text(encoding="utf-8"))

    gain_map = {(int(g["k"]), int(g["cluster"]), g["metric"]): g for g in gains}

    named = []
    for m in metrics:
        k, c = int(m["k"]), int(m["cluster"])
        g = gain_map[(k, c, "R@1")]
        prof = profile["clusters"].get(f"k{k}_c{c}", {})
        named.append({
            "k": k, "cluster": c, "cluster_name": names[(k, c)]["name"],
            "n": m["n"], "share": m["share"],
            "exp_01_R@1": m["exp_01_R@1"], "exp_04_R@1": m["exp_04_R@1"],
            "exp_01_R@5": m["exp_01_R@5"], "exp_01_MRR": m["exp_01_MRR"],
            "exp_01_nDCG@10": m["exp_01_nDCG@10"],
            "ltp_gain_R@1": g["gain"], "gain_ci_low": g["ci_low"], "gain_ci_high": g["ci_high"],
            "gain_significant": g["significant"],
            "hardneg_R@1": m["hardneg_R@1"], "hardneg_drop_R@1": m["hardneg_drop_R@1"],
            "UCR_L1": m["UCR_L1"], "UCR_L2": m["UCR_L2"],
            "difficulty_Hard_share": m.get("difficulty_Hard_share", ""),
            "median_duration_sec": prof.get("median_duration_sec", ""),
            "median_view_count": prof.get("median_view_count", ""),
            "top_title_terms": "；".join(
                f"{t['term']}×{t['ratio']}" for t in prof.get("top_title_terms", [])[:6]),
            "top_tag_terms": "；".join(
                f"{t['term']}×{t['ratio']}" for t in prof.get("top_tag_terms", [])[:6]),
        })
    write_csv(OUT_DIR / "video_cluster_named_metrics.csv", named)

    for a in assignments:
        for k in (3, 4):
            key = f"cluster_k{k}"
            if key in a:
                a[f"cluster_k{k}_name"] = names[(k, int(a[key]))]["name"]
    write_csv(OUT_DIR / "video_cluster_assignments_named.csv", assignments)

    write_markdown(OUT_DIR / "B4_論文表格.md", named, homo, quality, names)

    print(f"已輸出：{OUT_DIR / 'video_cluster_named_metrics.csv'}")
    print(f"已輸出：{OUT_DIR / 'video_cluster_assignments_named.csv'}")
    print(f"已輸出：{OUT_DIR / 'B4_論文表格.md'}")
    for k in (3, 4):
        print(f"  k={k}:")
        for r in [x for x in named if x["k"] == k]:
            print(f"    叢集{r['cluster']} {r['cluster_name']}（n={r['n']}，"
                  f"增益 {float(r['ltp_gain_R@1'])*100:+.2f}pp）")


def write_markdown(path, named, homo, quality, names):
    L = [f"# B4 短影音內容分層：論文表格（§4.7）\n",
         f"- 產生時間：{_dt.datetime.now().isoformat(timespec='seconds')}",
         "- 分層方式：影片 CLIP ViT-L/14 影像嵌入（12 幀平均、L2 正規化）之 k-means 語意叢集",
         "- 叢集名稱由人工檢視各叢集的判別性樣本與特徵詞富集比後命名\n"]

    for k in (3, 4):
        rows = [r for r in named if r["k"] == k]
        if not rows:
            continue
        q = next((x for x in quality if int(x["k"]) == k), {})
        L.append(f"\n## 表 4-x　k = {k} 的影片語意叢集分層結果"
                 f"（輪廓係數 {float(q.get('silhouette', 0)):.3f}）\n")
        L.append("| 叢集 | 名稱 | n | 占比 | Hybrid R@1 | No-LTP R@1 | LTP 增益 | "
                 "Hard-Neg R@1 | Hard-Neg 降幅 | UCR |")
        L.append("|---|---|---|---|---|---|---|---|---|---|")
        for r in rows:
            sig = "*" if str(r["gain_significant"]).lower() == "true" else ""
            L.append(
                f"| {r['cluster']} | {r['cluster_name']} | {r['n']} | "
                f"{float(r['share'])*100:.1f}% | {float(r['exp_01_R@1'])*100:.2f}% | "
                f"{float(r['exp_04_R@1'])*100:.2f}% | "
                f"{float(r['ltp_gain_R@1'])*100:+.2f}pp{sig} | "
                f"{float(r['hardneg_R@1'])*100:.2f}% | "
                f"{float(r['hardneg_drop_R@1'])*100:.2f}pp | "
                f"{float(r['UCR_L1'])*100:.2f}% |")
        L.append("\n註：`*` 表示 LTP 增益的配對拔靴 95% CI 不跨 0；"
                 "UCR 為子句層級（保守上界），母句校正後之值見 `video_cluster_named_metrics.csv`。\n")

        L.append(f"**k = {k} 各叢集的特徵詞（富集比 = 本叢集出現率 ÷ 其他叢集出現率）**\n")
        L.append("| 叢集 | 名稱 | 標題特徵詞 | 曲風／標籤特徵詞 | 長度中位數 | 觀看數中位數 |")
        L.append("|---|---|---|---|---|---|")
        for r in rows:
            dur = r["median_duration_sec"]
            vc = r["median_view_count"]
            L.append(f"| {r['cluster']} | {r['cluster_name']} | {r['top_title_terms']} | "
                     f"{r['top_tag_terms']} | "
                     f"{float(dur):.0f} 秒 | {float(vc):,.0f} |")

    L.append("\n## 增益同質性檢定\n")
    L.append("| k | 最高增益叢集 | 最低增益叢集 | 差距 | 事後 95% CI | 置換檢定 p | 結論 |")
    L.append("|---|---|---|---|---|---|---|")
    for h in homo:
        k = int(h["k"])
        hi_name = names[(k, int(h["cluster_max_gain"]))]["name"]
        lo_name = names[(k, int(h["cluster_min_gain"]))]["name"]
        p = float(h["permutation_p"])
        L.append(f"| {k} | {hi_name}（{float(h['gain_max'])*100:+.2f}pp）| "
                 f"{lo_name}（{float(h['gain_min'])*100:+.2f}pp）| "
                 f"{float(h['gain_range'])*100:+.2f}pp | "
                 f"[{float(h['ci_low'])*100:+.2f}, {float(h['ci_high'])*100:+.2f}] | "
                 f"{p:.4f} | {'增益因叢集而異' if p < 0.05 else '無證據支持增益隨叢集改變'} |")

    L.append("\n## 建議敘述（§4.7）\n")
    L.append("> 為檢驗合成偏好的效益是否僅限於特定影片類型，本研究以影片 CLIP 影像嵌入進行"
             "k-means 語意叢集，並由人工檢視各叢集的判別性樣本與特徵詞後命名。"
             "如表 4-x 所示，四個叢集分別對應主流流行舞曲音樂影片、粉絲二創與歌詞影片、"
             "樂團現場演出與搖滾金屬影片，以及嘻哈饒舌與街頭拍攝影片。"
             "在所有叢集中，Matched LTP 相對 No-LTP 的 R@1 增益皆達統計顯著"
             "（k = 4 時為 +9.85 至 +14.40 個百分點）。"
             "進一步以置換檢定檢驗增益是否因叢集而異（打亂叢集標籤並重算最大－最小增益全距），"
             "k = 3 與 k = 4 的 p 值分別為 .564 與 .174，均未達顯著。"
             "此結果顯示合成偏好的排序增益在不同影片語意叢集間具有同質性，"
             "並非僅在特定影片類型有效。\n")
    L.append("> **限制**：本資料集的「影片」即音樂自身之 YouTube 影片"
             "（video_id 與 target_music_id 相同），因此叢集反映的是音樂類型與影片形式"
             "（現場演出、粉絲二創、饒舌拍攝、流行舞曲），而非短影音創作者實際拍攝的"
             "旅遊、美食或知識教學等內容類型。分層結果不可外推至真實短影音創作情境。"
             "此外，叢集與候選音樂的聲學難度高度相關（各叢集 Hard 層占比差距達 32 個百分點），"
             "故各叢集基線 R@1 的差異主要反映聲學可辨識度，"
             "可歸因於內容分層的結論僅限於「增益具同質性」一項。\n")
    path.write_text("\n".join(L) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
